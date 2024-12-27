''' transform xlsx from jsy to jsons'''


import openpyxl
import json
import re
import sys


def open_workbook(input_file):
    try:
        workbook = openpyxl.load_workbook(input_file)
        if (len(workbook.sheetnames) != 4):
            raise Exception(f"open_workbook() -> {input_file} should have 4 sheets, but {len(workbook.sheetnames)} found")
        
        return workbook
    
    except FileNotFoundError:
        print(f"connections_to_json(): file {input_file} not found")
        sys.exit(-1)

def trans_external_port_coord(dir, x, y, index, version):
    port_index_array = [0,9,18,27,36,45,54,63,120,113,106,99,92,85,78,71]
    if version == 0:    # excel -> xl, output -> current
        dir_curr = "vert" if dir == 0 else "hori"
        x_curr = 9 - x
    else:               # excel -> current, output -> current
        dir_curr = "hori" if dir == 0 else "vert"
        x_curr = x
    
    return dir_curr, x_curr, y, port_index_array[index]

def bumps_to_json(input_file, output_file):
    try:
        workbook = open_workbook(input_file)

        topdies = {"muyan":{"pin_map":{}}}
        bump_pattern = r"bump_bank(\d+)_(\d+)"

        first_sheet = workbook.worksheets[0]
        for row in first_sheet.iter_rows(min_row=2, values_only=True):
            if row is None or not any(row):
                continue

            bump_name = str(row[0]) 
            signal_name = row[1]
            if signal_name is not None:
                match = re.match(bump_pattern, bump_name)
                if match:
                    num1, num2 = match.groups()
                    bump_index = int(num1)*64+int(num2)
                    assert 0 <= bump_index < 128, f"error with bump_index: {bump_index}"
                    topdies["muyan"]["pin_map"][signal_name] = bump_index
                else:
                    raise ValueError(f"bump name {bump_name} does not match pattern {bump_pattern}")
            
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(topdies, json_file, indent=4, ensure_ascii=False)
        
        print(f"pin_map 信息已保存到 {output_file}")

    except ValueError as e:
        print(f"bumps_to_json(): {e}")
        sys.exit(-1)
    except Exception as e:
        message = f"bumps_to_json() -> {e}"
        raise Exception(message)

def connections_to_json(input_file, output_file):
    try:
        workbook = open_workbook(input_file)
        
        connections = {}
        pattern = r"muyan_(\d+)_(.*)"

        sheet = workbook.worksheets[1]    
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row is None or not any(row):
                continue
            
            string1 = str(row[0])
            string2 = str(row[1])
            number_str = str(row[2])
            match1 = re.match(pattern, string1)
            match2 = re.match(pattern, string2)
            if match1:
                num, postfix = match1.groups()
                string1 = f"muyan_{num}.{postfix}"
            if match2:
                num, postfix = match2.groups()
                string2 = f"muyan_{num}.{postfix}"
            if number_str not in connections:
                connections[number_str] = [[string1, string2]]
            else:
                connections[number_str].append([string1, string2])
        
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(connections, json_file, indent=4, ensure_ascii=False)

        print(f"connections 信息已保存到 {output_file}")
    
    except Exception as e:
        message = f"connections_to_json() -> {e}"
        raise Exception(message)

def external_ports_to_json(input_file, output_file):
    try:
        workbook = open_workbook(input_file)

        external_ports = {}
        pattren = r"(\d+) (\d+) (\d+)"

        sheet = workbook.worksheets[2]
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row is None or not any(row):
                continue

            port_name = str(row[0])
            port_coord = str(row[1])
            port_array_index = str(row[2])
            match = re.match(pattren, port_coord)
            if match:
                dir, x, y = match.groups()
                d, row, col, index = trans_external_port_coord(int(dir), int(x), int(y), int(port_array_index), 0)
                external_ports[port_name] = {"coord": {"dir": d, "row": row, "col":col, "index":index}}
            else:
                raise ValueError(f"fail to match {port_coord} with pattern {pattren}")
        
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(external_ports, json_file, indent=4, ensure_ascii=False)

        print(f"external ports 信息已保存到 {output_file}")

    except ValueError as e:
        print(f"external_ports_to_json(): {e}")
        sys.exit(-1)
    except Exception as e:
        message = f"external_ports_to_json() -> {e}"
        raise Exception(message)
    
def interposer_json(output_file):
    try:
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump({}, json_file, indent=4, ensure_ascii=False)

        print(f"interposer 信息已保存到 {output_file}")

    except Exception as e:
        message = f"interposer_json() -> {e}"
        raise Exception(message)

def topdie_insts_json(output_file, name, number):
    try:
        topdie_insts = {}
        for n in range(number):
            row = n // 4
            col = n % 4
            topdie_insts[name+"_"+str(n)] = {"topdie": name, "coord": {"row": row, "col": col}}

        with open(output_file, 'w', encoding='utf-8') as json_file:
                json.dump(topdie_insts, json_file, indent=4, ensure_ascii=False)

        print(f"topdie insts 信息已保存到 {output_file}")
    
    except Exception as e:
        message = f"topdie_insts_json() -> {e}"
        raise Exception(message)

def reigster_adder_json(output_file):
    try:
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump({}, json_file, indent=4, ensure_ascii=False)

        print(f"reigster 信息已保存到 {output_file}")

    except Exception as e:
        message = f"reigster_adder_json() -> {e}"
        raise Exception(message)

def config_json(output_file):
    try:
        config = {
            "interposer" : "interposer.json",
            "topdies" : "topdies.json",
            "topdie_insts" : "topdie_insts.json",
            "external_ports" : "external_ports.json",
            "connections" : "connections.json",
            "reigster_adder" : "reigster_adder.json"
        }
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(config, json_file, indent=4, ensure_ascii=False)

        print(f"config 信息已保存到 {output_file}")

    except Exception as e:
        message = f"config_json() -> {e}"
        raise Exception(message)


if __name__ == "__main__":
    input_xlsx = "config/case6.xlsx"   

    connections_to_json(input_xlsx, "config/connections.json")
    external_ports_to_json(input_xlsx, "config/external_ports.json")
    bumps_to_json(input_xlsx, "config/topdies.json")
    interposer_json("config/interposer.json")
    topdie_insts_json("config/topdie_insts.json", "muyan", 16)
    reigster_adder_json("config/reigster_adder.json")
    config_json("config/config.json")
